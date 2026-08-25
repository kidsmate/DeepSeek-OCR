# -*- coding: utf-8 -*-
"""
X-Ray工装空洞分析全流程工具（批量版）
功能：批量处理任务令（MD/MN 11位，DD 12位），Chip/clip标准化（无空格），合并，匹配，可视化
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
import re, os, shutil, datetime, time, traceback
from pathlib import Path
from PIL import Image as PILImage
import threading
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.patches import FancyBboxPatch, Patch

class AppConfig:
    BG_COLOR = "#ffffff"
    FG_COLOR = "#000000"
    LOG_FONT = ("微软雅黑", 9)
    NORMAL_FONT = ("微软雅黑", 9)
    BOLD_FONT = ("微软雅黑", 10, "bold")
    TASK_FOLDER_PATTERN = re.compile(r"^(?:MD|MN)[A-Za-z0-9]{9}$|^DD[A-Za-z0-9]{10}$", re.I)
    # 数据文件夹规则：
    #   019开头BP：15位以0开头（0 + 14位BP码）
    #   599开头BP：14位以599开头（BP码本身，无前导0）
    DATA_FOLDER_PATTERN = r"^(0\d{14}|599\d{11})(NG\d*)?(_\d+)?$"
    IMAGE_SUFFIX = (".png", ".jpg", ".jpeg", ".bmp")
    TARGET_IMG_SIZE = (265, 265)
    DIR_IMG_EXPORT = "1_提取X-Ray图片目录"
    DIR_BARCODE = "2_条码清单"
    DIR_CSV_EXCEL = "3_CSV转Excel"
    DIR_MERGE_VOID = "5_合并空洞数据"
    DIR_AMB_CHIP = "6_AMB_Chip提取结果"
    DIR_TEMP_IMG = "temp临时图片"
    DIR_BATCH_SUMMARY = "图片匹配结果汇总"
    DIR_MERGE_SUMMARY = "合并后的图片匹配结果"
    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HEADER_FONT = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")
    CONTENT_FONT = Font(name="微软雅黑", size=9)
    FOOTER_FONT = Font(name="微软雅黑", size=9)
plt.rcParams["font.sans-serif"] = ["Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False

class LogUtil:
    _widget = None
    _root = None
    @classmethod
    def bind(cls, widget, root_win):
        cls._widget = widget
        cls._root = root_win
    @classmethod
    def append(cls, msg):
        ts = datetime.datetime.now().strftime("[%H:%M:%S]")
        line = f"{ts} {msg}\n"
        def _write():
            if cls._widget is None:
                return
            cls._widget.insert(tk.END, line)
            cls._widget.see(tk.END)
        if cls._root:
            cls._root.after(0, _write)

class PathUtil:
    @staticmethod
    def clean(s):
        return str(s).strip().strip('"').strip("'")
    @staticmethod
    def find_all_csv(root):
        csv_list = []
        for r, _, files in os.walk(root):
            for f in files:
                if f.lower().endswith(".csv"):
                    csv_list.append(os.path.join(r, f))
        return csv_list
    @staticmethod
    def find_bp_excel(root):
        if not os.path.isdir(root):
            return None
        best = None
        best_score = -1
        candidates = []
        for dirpath, _, filenames in os.walk(root):
            for f in filenames:
                if not f.lower().endswith((".xlsx", ".xls")):
                    continue
                full = os.path.join(dirpath, f)
                fname_lower = f.lower()
                score = 0
                if "每日加工记录" in fname_lower:
                    score += 200
                if "bp" in fname_lower:
                    score += 100
                if "工装" in fname_lower and "位号" in fname_lower:
                    score += 50
                if "空洞" in fname_lower or "关系" in fname_lower:
                    score -= 200
                if score > 0:
                    candidates.append((score, full))
                if score > best_score:
                    best_score = score
                    best = full
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            LogUtil.append(f"  │ BP候选文件（前3个）：")
            for i, (sc, p) in enumerate(candidates[:3]):
                LogUtil.append(f"  │   {i+1}. {os.path.basename(p)} (分数:{sc})")
        if best:
            LogUtil.append(f"  │ 选中BP文件：{os.path.basename(best)}")
        return best
    @staticmethod
    def find_all_excel(root):
        res = []
        if not os.path.isdir(root):
            return res
        for dirpath, _, filenames in os.walk(root):
            for fname in filenames:
                if fname.lower().endswith(".xlsx") and not fname.startswith("~$"):
                    res.append(os.path.join(dirpath, fname))
        return res

class ChipClipUtil:
    @staticmethod
    def normalize_cell_value(value, target_prefix):
        if value is None:
            return None
        text = str(value).strip()
        if not text or text.upper() == "AMB":
            return text
        if text.lower().startswith(target_prefix.lower()):
            return text
        cleaned = re.sub(r"^(?:chip|clip)\s*", "", text, flags=re.I).strip()
        if cleaned == "":
            return None
        return f"{target_prefix}{cleaned}"

class SingleTaskProcessor:
    def __init__(self, task_folder_path):
        self.task_path = PathUtil.clean(task_folder_path)
        self.task_name = os.path.basename(self.task_path)
        parent = os.path.dirname(self.task_path)
        self.root_dir = parent
        self.out_base = os.path.join(parent, f"{self.task_name}_自动输出")
        self.tmp_img_dir = os.path.join(self.out_base, AppConfig.DIR_TEMP_IMG)
        self.csv_files = PathUtil.find_all_csv(self.task_path)
        self.bp_excel = PathUtil.find_bp_excel(self.root_dir)
        self.result_excel = ""
    def is_processed(self):
        if not os.path.isdir(self.out_base):
            return False
        for f in os.listdir(self.out_base):
            if f.endswith("_图片匹配汇总结果.xlsx"):
                self.result_excel = os.path.join(self.out_base, f)
                return True
        return False
    def _extract_folders(self):
        LogUtil.append("  ┌ 开始提取有效数据文件夹")
        out_img = os.path.join(self.out_base, AppConfig.DIR_IMG_EXPORT)
        out_bc = os.path.join(self.out_base, AppConfig.DIR_BARCODE)
        os.makedirs(out_img, exist_ok=True)
        os.makedirs(out_bc, exist_ok=True)
        bc_list = []
        valid = []
        pat = re.compile(AppConfig.DATA_FOLDER_PATTERN, re.I)
        for root, dirs, _ in os.walk(self.task_path):
            for d in dirs:
                m = pat.fullmatch(d)
                if m:
                    full_p = os.path.join(root, d)
                    code15 = m.group(1)
                    ng = m.group(2) or ""
                    key = code15 + ng
                    # 019开头：15位以0开头，去掉前导0得到14位BP码
                    # 599开头：14位以599开头，本身即为BP码，无需去掉前导字符
                    display = (code15[1:] if len(code15) == 15 else code15) + ng
                    valid.append((full_p, d, key, display))
        LogUtil.append(f"  │ 扫描完成，共发现 {len(valid)} 个候选数据文件夹")
        key_set = set()
        for p, name, key, disp in valid:
            if key in key_set:
                continue
            key_set.add(key)
            bc_list.append(disp)
            dst = os.path.join(out_img, name)
            n = 1
            while os.path.exists(dst):
                dst = os.path.join(out_img, f"{name}_{n}")
                n += 1
            try:
                shutil.copytree(p, dst)
            except Exception as e:
                LogUtil.append(f"  │ ⚠ 复制数据文件夹失败：{name} -> {e}")
        with open(os.path.join(out_bc, "条码清单.txt"), "w", encoding="utf-8") as f:
            f.write(",".join(bc_list))
        LogUtil.append(f"  └ 去重完成，有效数据文件夹 {len(valid)} 个，有效条码 {len(bc_list)} 条")
        return bc_list, out_img
    def _csv_to_excel(self):
        LogUtil.append("  ┌ 开始CSV转Excel并处理Chip/Clip（基于AMB位置）")
        out_dir = os.path.join(self.out_base, AppConfig.DIR_CSV_EXCEL)
        os.makedirs(out_dir, exist_ok=True)
        if not self.csv_files:
            LogUtil.append("  └ 未找到CSV文件，跳过")
            return out_dir
        LogUtil.append(f"  │ 共 {len(self.csv_files)} 个CSV文件待处理")
        success = 0
        for csv_p in self.csv_files:
            fname = os.path.basename(csv_p)
            try:
                df = pd.read_csv(csv_p, dtype=str)
                if df.shape[1] >= 9:
                    col_i = df.columns[8]
                    amb_indices = df[df[col_i].str.upper() == "AMB"].index.tolist()
                    for amb_idx in amb_indices:
                        start = max(0, amb_idx - 12)
                        for r in range(start, amb_idx):
                            if r < 0 or r >= len(df):
                                continue
                            val = df.at[r, col_i]
                            if pd.isna(val):
                                continue
                            text = str(val).strip()
                            if text.upper() == "AMB":
                                continue
                            if not text.lower().startswith("chip"):
                                new_val = ChipClipUtil.normalize_cell_value(text, "Chip")
                                if new_val is not None:
                                    df.at[r, col_i] = new_val
                        start2 = max(0, amb_idx - 42)
                        end2 = amb_idx - 13
                        for r in range(start2, end2 + 1):
                            if r < 0 or r >= len(df):
                                continue
                            val = df.at[r, col_i]
                            if pd.isna(val):
                                continue
                            text = str(val).strip()
                            if text.upper() == "AMB":
                                continue
                            if not text.lower().startswith("clip"):
                                new_val = ChipClipUtil.normalize_cell_value(text, "Clip")
                                if new_val is not None:
                                    df.at[r, col_i] = new_val
                xlsx = os.path.join(out_dir, os.path.splitext(fname)[0] + ".xlsx")
                df.to_excel(xlsx, index=False)
                wb = load_workbook(xlsx)
                ws = wb.active
                for col in ["C", "D", "E", "F"]:
                    ws.column_dimensions[col].width = 16
                for r in range(2, ws.max_row + 1):
                    for c in ["C", "D", "E", "F"]:
                        cell = ws[f"{c}{r}"]
                        if cell.value is None:
                            continue
                        try:
                            cell.value = int(float(str(cell.value).strip()))
                            cell.number_format = numbers.FORMAT_NUMBER
                        except:
                            pass
                wb.save(xlsx)
                wb.close()
                success += 1
            except Exception as e:
                LogUtil.append(f"  │ ⚠ {fname} 转换失败：{e}")
        LogUtil.append(f"  └ 转换完成，成功 {success} 个，输出目录：{out_dir}")
        return out_dir
    def _merge_void(self, src_dir):
        LogUtil.append("  ┌ 开始空洞数据合并与百分比转数值计算")
        save_dir = os.path.join(self.out_base, AppConfig.DIR_MERGE_VOID)
        os.makedirs(save_dir, exist_ok=True)
        files = list(Path(src_dir).glob("*.xlsx"))
        files = [f for f in files if not f.name.startswith("~$")]
        if not files:
            LogUtil.append("  └ 无文件可合并，跳过")
            return ""
        LogUtil.append(f"  │ 待合并文件 {len(files)} 个")
        dfs = []
        for f in files:
            try:
                dfs.append(pd.read_excel(f, dtype=str))
            except Exception as e:
                LogUtil.append(f"  │ ⚠ 读取失败：{f.name} -> {e}")
        if not dfs:
            return ""
        merge_df = pd.concat(dfs, ignore_index=True)
        out_name = f"合并Excel数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_p = os.path.join(save_dir, out_name)
        wb = Workbook()
        ws = wb.active
        ws.title = "合并数据"
        for r in dataframe_to_rows(merge_df, index=False, header=True):
            ws.append(r)
        LogUtil.append(f"  │ 合并后共 {len(merge_df)} 行数据")
        ws.insert_cols(12)
        ws.cell(1, 12, "Single Void 数值")
        ws.insert_cols(14)
        ws.cell(1, 14, "Total Void 数值")
        LogUtil.append("  │ 开始百分比转数值计算")
        for r in range(2, ws.max_row + 1):
            for src_col, dst_col in [(11, 12), (13, 14)]:
                v = ws.cell(r, src_col).value
                num = None
                if v:
                    s = str(v).strip()
                    if s.endswith("%"):
                        try:
                            num = round(float(s[:-1]) / 100, 5)
                        except:
                            pass
                    else:
                        try:
                            num = round(float(s), 5)
                        except:
                            pass
                cell = ws.cell(r, dst_col, num)
                cell.number_format = "0.00000"
        wb.save(out_p)
        wb.close()
        LogUtil.append(f"  └ 合并完成，输出文件：{out_name}")
        return out_p
    def _extract_amb_chip(self, merge_dir):
        LogUtil.append("  ┌ 开始AMB/Chip分类提取（按类型严格分离）")
        out_dir = os.path.join(self.out_base, AppConfig.DIR_AMB_CHIP)
        os.makedirs(out_dir, exist_ok=True)
        amb_dir = os.path.join(out_dir, "AMB_数据汇总")
        joint_dir = os.path.join(out_dir, "12_Chip 数据汇总")
        chip_dir = os.path.join(out_dir, "30_Clip 数据汇总")
        os.makedirs(amb_dir, exist_ok=True)
        os.makedirs(joint_dir, exist_ok=True)
        os.makedirs(chip_dir, exist_ok=True)
        files = PathUtil.find_all_excel(merge_dir)
        if not files:
            LogUtil.append("  └ 未找到合并源文件，跳过")
            return
        src = files[0]
        LogUtil.append(f"  │ 读取源文件：{os.path.basename(src)}")
        wb_src = load_workbook(src, data_only=True)
        ws_src = wb_src.active
        max_col, max_row = ws_src.max_column, ws_src.max_row
        LogUtil.append(f"  │ 源文件共 {max_row - 1} 行数据，{max_col} 列")
        type_col = 9
        LogUtil.append(f"  │ 分类依据：I列（第9列）的内容前缀")
        wb_amb, wb_joint, wb_chip = Workbook(), Workbook(), Workbook()
        ws_amb, ws_joint, ws_chip = wb_amb.active, wb_joint.active, wb_chip.active
        ws_amb.title, ws_joint.title, ws_chip.title = "AMB数据", "12_Chip数据", "30_Clip数据"
        for c in range(1, max_col + 1):
            header = ws_src.cell(1, c).value
            ws_amb.cell(1, c, header)
            ws_joint.cell(1, c, header)
            ws_chip.cell(1, c, header)
        row_amb = row_joint = row_chip = 2
        for r in range(2, max_row + 1):
            type_val = str(ws_src.cell(r, type_col).value).strip() if ws_src.cell(r, type_col).value else ""
            type_low = type_val.lower()
            if type_val.upper() == "AMB":
                for c in range(1, max_col + 1):
                    ws_amb.cell(row_amb, c, ws_src.cell(r, c).value)
                row_amb += 1
            elif type_low.startswith("chip"):
                for c in range(1, max_col + 1):
                    ws_joint.cell(row_joint, c, ws_src.cell(r, c).value)
                row_joint += 1
            elif type_low.startswith("clip"):
                for c in range(1, max_col + 1):
                    ws_chip.cell(row_chip, c, ws_src.cell(r, c).value)
                row_chip += 1
        wb_amb.save(os.path.join(amb_dir, "AMB数据.xlsx"))
        wb_joint.save(os.path.join(joint_dir, "12_Chip数据.xlsx"))
        wb_chip.save(os.path.join(chip_dir, "30_Clip数据.xlsx"))
        wb_src.close()
        wb_amb.close()
        wb_joint.close()
        wb_chip.close()
        LogUtil.append(f"  │ AMB数据：{row_amb - 2} 行")
        LogUtil.append(f"  │ 12_Chip数据：{row_joint - 2} 行")
        LogUtil.append(f"  │ 30_Clip数据：{row_chip - 2} 行")
        LogUtil.append(f"  └ 分类提取完成，输出根目录：{out_dir}")
    def _generate_match_excel(self, bc_list, img_dir):
        LogUtil.append("  ┌ 开始生成图片匹配汇总表")
        wb = Workbook()
        ws = wb.active
        ws.title = "匹配结果"
        headers = ["条码信息", "M1位置X-Ray图片", "M2位置X-Ray图片", "M3位置X-Ray图片", "匹配状态", "匹配总数", "结果", "空洞总数", "U相空洞数", "V相空洞数", "W相空洞数", "模组位置", "工装编号"]
        for i, h in enumerate(headers, 1):
            ws.cell(1, i, h)
        col_w = {"A": 8, "B": 38, "C": 38, "D": 38, "G": 7, "H": 8, "I": 10, "J": 10, "K": 10, "L": 8, "M": 10}
        for c, w in col_w.items():
            ws.column_dimensions[c].width = w
        row_ptr = 2
        os.makedirs(self.tmp_img_dir, exist_ok=True)
        def find_folder(code):
            for r, ds, _ in os.walk(img_dir):
                for d in ds:
                    if code.lower() in d.lower():
                        return os.path.join(r, d)
            return None
        match_success = 0
        for bc in bc_list:
            ng = re.search(r"NG\d*$", bc, re.I)
            pure = bc[:-len(ng.group())] if ng else bc
            ws[f"A{row_ptr}"] = bc
            ws[f"G{row_ptr}"] = "NG" if ng else "OK"
            folder = find_folder(pure)
            imgs = []
            if folder:
                for f in os.listdir(folder):
                    fp = os.path.join(folder, f)
                    if os.path.isfile(fp) and f.lower().endswith(AppConfig.IMAGE_SUFFIX):
                        if pure.lower() in f.lower() and os.path.splitext(f)[0].lower().endswith("dbc"):
                            imgs.append(fp)
            imgs = imgs[:3]
            if len(imgs) == 3:
                match_success += 1
            for i, col in enumerate(["B", "C", "D"]):
                if i >= len(imgs):
                    continue
                tmp = os.path.join(self.tmp_img_dir, f"{pure}_img{i + 1}{os.path.splitext(imgs[i])[1]}")
                shutil.copy2(imgs[i], tmp)
                try:
                    with PILImage.open(tmp) as im:
                        im = im.resize(AppConfig.TARGET_IMG_SIZE, PILImage.Resampling.LANCZOS)
                        im.save(tmp, dpi=(96, 96))
                except Exception as e:
                    LogUtil.append(f"  │ ⚠ 图片处理失败：{tmp} -> {e}")
                    continue
                img_obj = Image(tmp)
                img_obj.width, img_obj.height = AppConfig.TARGET_IMG_SIZE
                ws.add_image(img_obj, f"{col}{row_ptr}")
            status = ["无匹配图片", "部分匹配成功", "全部匹配成功"][min(len(imgs), 2)]
            ws[f"E{row_ptr}"] = status
            ws[f"F{row_ptr}"] = len(imgs)
            ws[f"H{row_ptr}"] = f"=SUM(I{row_ptr}:K{row_ptr})"
            ws.row_dimensions[row_ptr].height = 205
            sub = row_ptr + 1
            ws[f"A{sub}"] = "X-Ray图片源文件"
            for i, col in enumerate(["B", "C", "D"]):
                if i < len(imgs):
                    ws[f"{col}{sub}"] = os.path.basename(imgs[i])
            ws.row_dimensions[sub].height = 47
            row_ptr += 2
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = AppConfig.HEADER_FONT, AppConfig.HEADER_FILL, AppConfig.CENTER_ALIGN
        for r in ws.iter_rows(min_row=2, max_row=row_ptr - 1, min_col=1, max_col=13):
            for cell in r:
                cell.alignment, cell.font = AppConfig.CENTER_ALIGN, AppConfig.CONTENT_FONT
        for r in ws.iter_rows():
            if r[0].value == "X-Ray图片源文件":
                for cell in r:
                    cell.font = AppConfig.FOOTER_FONT
        save_name = f"{self.task_name}_图片匹配汇总结果.xlsx"
        out_p = os.path.join(self.out_base, save_name)
        wb.save(out_p)
        wb.close()
        time.sleep(0.5)
        shutil.rmtree(self.tmp_img_dir, ignore_errors=True)
        LogUtil.append(f"  │ 处理条码 {len(bc_list)} 条，完全匹配 {match_success} 条")
        LogUtil.append(f"  └ 汇总表生成完成：{save_name}")
        return out_p
    def _bp_fill(self, excel_p):
        # 根据最新需求：只处理偶数行，BP码从D列取，工装编号取C列，模组位置取H列
        if not self.bp_excel or not os.path.exists(self.bp_excel):
            LogUtil.append("  ℹ 未找到BP台账文件，跳过回填")
            return
        LogUtil.append(f"  │ 使用BP文件：{os.path.basename(self.bp_excel)}")
        try:
            df_bp = pd.read_excel(self.bp_excel, dtype=str)
            # 固定列索引：D列是BP码（索引3），C列是工装编号（索引2），H列是模组位置（索引7）
            bp_col_idx = 3  # D列
            gz_col_idx = 2  # C列
            pos_col_idx = 7 # H列
            # 检查列数是否足够
            if len(df_bp.columns) <= max(bp_col_idx, gz_col_idx, pos_col_idx):
                LogUtil.append(f"  ℹ BP表列数不足，跳过回填")
                return
            bp_map = {}
            for idx, row in df_bp.iterrows():
                code = str(row.iloc[bp_col_idx]).strip()
                if re.fullmatch(r"\d{14}", code):
                    gz_val = str(row.iloc[gz_col_idx]).strip()
                    pos_val = str(row.iloc[pos_col_idx]).strip()
                    bp_map[code] = (gz_val, pos_val)
            LogUtil.append(f"  │ BP表有效映射数：{len(bp_map)} 条")
            if not bp_map:
                LogUtil.append("  ℹ BP表中无有效的14位BP码，跳过回填")
                return
            # 读取汇总表，只处理偶数行（从第2行开始，步长2）
            wb = load_workbook(excel_p)
            ws = wb.active
            match_cnt = 0
            # 先收集一些示例条码（从偶数行）
            sample_codes = []
            max_row = ws.max_row
            for row in range(2, min(max_row, 12), 2):
                a = str(ws[f"A{row}"].value).strip() if ws[f"A{row}"].value else ""
                m = re.search(r"\d{14}", a)
                if m:
                    sample_codes.append(m.group())
            if sample_codes:
                LogUtil.append(f"  │ 汇总表条码示例（偶数行前几个14位码）：{sample_codes[:5]}")
            # 实际回填
            for row in range(2, max_row + 1, 2):  # 只遍历偶数行
                a = str(ws[f"A{row}"].value).strip() if ws[f"A{row}"].value else ""
                m = re.search(r"\d{14}", a)
                if not m:
                    continue
                bc14 = m.group()
                if bc14 in bp_map:
                    gz_val, pos_val = bp_map[bc14]
                    ws[f"L{row}"] = pos_val   # L列 = 模组位置（H列）
                    ws[f"M{row}"] = gz_val    # M列 = 工装编号（C列）
                    # 设置格式
                    ws[f"L{row}"].number_format = "0"
                    ws[f"M{row}"].number_format = "0"
                    match_cnt += 1
            wb.save(excel_p)
            wb.close()
            if match_cnt > 0:
                LogUtil.append(f"  │ BP回填成功：匹配 {match_cnt} 条记录（偶数行）")
            else:
                LogUtil.append(f"  ℹ BP表有 {len(bp_map)} 条映射，但在汇总表偶数行中未匹配到任何条码")
        except Exception as e:
            LogUtil.append(f"  ⚠ BP回填异常：{e}")
            LogUtil.append(traceback.format_exc())
    def run(self):
        try:
            bc_list, img_dir = self._extract_folders()
            if not bc_list:
                LogUtil.append("  ⚠ 无有效条码数据，终止处理")
                return ""
            csv_dir = self._csv_to_excel()
            merge_dir = os.path.join(self.out_base, AppConfig.DIR_MERGE_VOID)
            self._merge_void(csv_dir)
            self._extract_amb_chip(merge_dir)
            match_excel = self._generate_match_excel(bc_list, img_dir)
            self._bp_fill(match_excel)
            self.result_excel = match_excel
            return match_excel
        except Exception as e:
            LogUtil.append(f"  ❌ 处理异常：{e}")
            LogUtil.append(traceback.format_exc())
            return ""

class BatchTaskRunner:
    def __init__(self, root_path):
        self.root_path = PathUtil.clean(root_path)
        self.task_folders = []
        self.result_files = []
    def scan_tasks(self):
        pat = AppConfig.TASK_FOLDER_PATTERN
        LogUtil.append("扫描目录结构：总根目录 → 下一层任务令文件夹")
        LogUtil.append("任务令规则：MD/MN开头11位；DD开头12位")
        if not os.path.isdir(self.root_path):
            LogUtil.append("❌ 总根目录不存在")
            return 0
        try:
            entries = os.listdir(self.root_path)
        except Exception as e:
            LogUtil.append(f"❌ 无法读取总根目录：{e}")
            return 0
        for entry in entries:
            full_p = os.path.join(self.root_path, entry)
            if os.path.isdir(full_p) and pat.fullmatch(entry):
                self.task_folders.append(full_p)
        self.task_folders.sort(key=lambda p: os.path.basename(p).upper())
        LogUtil.append(f"扫描完成：符合条件的任务令 {len(self.task_folders)} 个")
        for i, task in enumerate(self.task_folders, 1):
            LogUtil.append(f"  [{i}] {os.path.basename(task)}")
        return len(self.task_folders)
    def run_all(self, finish_cb):
        total = len(self.task_folders)
        skip = done = 0
        new_result_files, skipped_files = [], []
        self.result_files = []
        for idx, task_p in enumerate(self.task_folders, 1):
            task_name = os.path.basename(task_p)
            LogUtil.append(f"\n{'=' * 50}")
            LogUtil.append(f"[{idx}/{total}] 当前任务令：{task_name}")
            processor = SingleTaskProcessor(task_p)
            if processor.is_processed():
                LogUtil.append("ℹ 该任务令已处理过，自动跳过处理流程")
                self.result_files.append(processor.result_excel)
                skipped_files.append(processor.result_excel)
                skip += 1
                continue
            LogUtil.append("▶ 开始执行完整流水线...")
            result = processor.run()
            if result:
                self.result_files.append(result)
                new_result_files.append(result)
                done += 1
                LogUtil.append(f"✅ 任务令 {task_name} 处理完成")
            else:
                LogUtil.append(f"⚠ 任务令 {task_name} 未生成有效结果")
        LogUtil.append(f"\n{'=' * 50}")
        LogUtil.append("批量处理全部完成")
        LogUtil.append(f"总计：{total} 个任务令，新增处理 {done} 个，自动跳过 {skip} 个")
        LogUtil.append(f"有效汇总表共 {len(self.result_files)} 份")
        if self.result_files:
            summary_dir = os.path.join(self.root_path, AppConfig.DIR_BATCH_SUMMARY)
            os.makedirs(summary_dir, exist_ok=True)
            LogUtil.append("\n开始统一归档汇总表...")
            LogUtil.append(f"  目标文件夹：{summary_dir}")
            for fp in new_result_files:
                try:
                    shutil.copy2(fp, os.path.join(summary_dir, os.path.basename(fp)))
                except Exception as e:
                    LogUtil.append(f"  ⚠ 归档失败：{fp} -> {e}")
            if new_result_files:
                LogUtil.append(f"✅ 新生成汇总表归档完成，共复制 {len(new_result_files)} 份")
            for fp in skipped_files:
                LogUtil.append(f"  ℹ {os.path.basename(fp)}：图片匹配结果表无变化，无需拷贝处理")
            if skipped_files:
                LogUtil.append(f"  共 {len(skipped_files)} 份汇总表无需重复拷贝")
        finish_cb(self.result_files)

class SummaryMerger:
    def __init__(self, output_root_dir):
        self.output_root = PathUtil.clean(output_root_dir)
        self.merge_dir = os.path.join(self.output_root, AppConfig.DIR_MERGE_SUMMARY)
        self.merged_file = ""
    def run(self, file_list):
        LogUtil.append("=" * 50)
        LogUtil.append("【多表合并】开始执行")
        try:
            if not file_list:
                LogUtil.append("❌ 未找到任何汇总结果文件")
                return ""
            LogUtil.append(f"待合并汇总表共 {len(file_list)} 份")
            for i, f in enumerate(file_list, 1):
                LogUtil.append(f"  [{i}] {os.path.basename(f)}")
            LogUtil.append("步骤1/5：开始读取并合并所有数据...")
            all_dfs = []
            for i, fp in enumerate(file_list):
                try:
                    df = pd.read_excel(fp, dtype=str, header=0)
                    all_dfs.append(df)
                    LogUtil.append(f"  读取第{i + 1}份：{os.path.basename(fp)}，{len(df)} 行 {len(df.columns)} 列")
                except Exception as e:
                    LogUtil.append(f"  ⚠ 读取失败 {os.path.basename(fp)}：{e}")
            if not all_dfs:
                LogUtil.append("❌ 无有效数据可合并")
                return ""
            LogUtil.append("  正在按列对齐拼接...")
            merge_df = pd.concat(all_dfs, ignore_index=True, sort=False)
            os.makedirs(self.merge_dir, exist_ok=True)
            out_name = f"多任务令汇总总表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_p = os.path.join(self.merge_dir, out_name)
            merge_df.to_excel(out_p, index=False)
            LogUtil.append(f"  数据合并完成，总行数：{len(merge_df)}，总列数：{len(merge_df.columns)}")
            LogUtil.append("步骤2/5：开始设置列宽...")
            wb_fmt = load_workbook(out_p)
            ws_fmt = wb_fmt.active
            max_row = ws_fmt.max_row
            widths = {"A": 16, "B": 23, "C": 23, "D": 23, "E": 14, "F": 9, "G": 6, "H": 10, "I": 12, "J": 12, "K": 12, "L": 10, "M": 23}
            for col, width in widths.items():
                ws_fmt.column_dimensions[col].width = width
            LogUtil.append("  列宽设置完成")
            LogUtil.append("步骤3/5：开始设置字体、对齐与自动换行...")
            header_font, content_font = Font(name="微软雅黑", size=9, bold=True), Font(name="微软雅黑", size=9, bold=False)
            center_align, wrap_align = Alignment(horizontal="center", vertical="center"), Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cell in ws_fmt[1]:
                cell.font, cell.alignment = header_font, center_align
            for row in ws_fmt.iter_rows(min_row=2, max_row=max_row):
                for cell in row:
                    cell.font = content_font
                    cell.alignment = wrap_align if cell.column_letter in ("B", "C", "D") else center_align
            LogUtil.append("  字体、对齐与自动换行设置完成")
            LogUtil.append("步骤4/5：开始设置行高...")
            odd_count = 0
            for row_num in range(3, max_row + 1):
                if row_num % 2 == 1:
                    ws_fmt.row_dimensions[row_num].height = 55
                    odd_count += 1
            LogUtil.append(f"  行高设置完成，共设置 {odd_count} 行")
            LogUtil.append("步骤5/5：偶数行H列SUM + I/J/K/L数值化...")
            formula_count = 0
            for row_num in range(2, max_row + 1):
                if row_num % 2 == 0:
                    cell = ws_fmt.cell(row=row_num, column=8)
                    cell.value = f"=SUM(I{row_num}:K{row_num})"
                    cell.number_format = numbers.FORMAT_NUMBER
                    formula_count += 1
            LogUtil.append(f"  H列公式设置完成，共 {formula_count} 行")
            num_cols = [9, 10, 11, 12]
            convert_success = convert_fail = 0
            for row_num in range(2, max_row + 1):
                if row_num % 2 == 1:
                    continue
                for col_idx in num_cols:
                    cell = ws_fmt.cell(row=row_num, column=col_idx)
                    val = cell.value
                    if val is None or str(val).strip() == "":
                        cell.value = 0
                        convert_success += 1
                        continue
                    try:
                        s_val = str(val).strip()
                        cell.value = float(s_val) if "." in s_val else int(s_val)
                        convert_success += 1
                    except:
                        convert_fail += 1
            LogUtil.append(f"  I/J/K/L数值转换完成：成功 {convert_success} 个，失败 {convert_fail} 个")
            wb_fmt.save(out_p)
            wb_fmt.close()
            self.merged_file = out_p
            LogUtil.append("✅ 多表合并全部完成")
            LogUtil.append(f"  输出文件：{out_name}")
            LogUtil.append(f"  保存路径：{out_p}")
            return out_p
        except Exception as e:
            LogUtil.append(f"❌ 合并异常：{e}")
            LogUtil.append(traceback.format_exc())
            return ""

class VoidMatcher:
    def __init__(self, summary_file, rel_file):
        self.summary_path = PathUtil.clean(summary_file)
        self.rel_path = PathUtil.clean(rel_file)
        self.result_path = ""
        self.gz_total = {}
        self.position_total = []
    def run(self):
        LogUtil.append("=" * 50)
        LogUtil.append("【工装匹配】开始执行")
        if not os.path.exists(self.summary_path) or not os.path.exists(self.rel_path):
            LogUtil.append("❌ 文件路径无效")
            return False
        try:
            pat_full = re.compile(r"^GZ.{18}$")
            LogUtil.append("步骤1/6：打开工装及在炉位置记录表...")
            wb_rel = load_workbook(self.rel_path, data_only=True)
            ws_rel = wb_rel.active
            max_row = max(ws_rel.max_row, 660)
            LogUtil.append(f"  工作表加载成功，有效数据范围：{max_row} 行，{ws_rel.max_column} 列")
            LogUtil.append(f"  工作表名称：{ws_rel.title}")
            LogUtil.append("步骤2/6：解析工装编号，构建双重映射关系...")
            full_map, prefix14_map = {}, {}
            empty_count = 0
            for r in range(1, max_row + 1):
                val = ws_rel.cell(r, 1).value
                if not val:
                    empty_count += 1
                    continue
                code = re.sub(r"\s+", "", str(val)).split("NG")[0]
                if pat_full.match(code):
                    full_map[code] = r
                if len(code) >= 14:
                    prefix14_map[code[:14]] = r
            wb_rel.close()
            LogUtil.append(f"  共扫描 {max_row} 行，空行 {empty_count} 行，完整19位工装号 {len(full_map)} 条，14位前缀 {len(prefix14_map)} 条")
            LogUtil.append("步骤3/6：读取合并后的汇总总表...")
            df = pd.read_excel(self.summary_path)
            df.columns = [str(c).strip() for c in df.columns]
            LogUtil.append(f"  汇总表读取成功，共 {len(df)} 条数据，{len(df.columns)} 列字段")
            need = ["工装编号", "U相空洞数", "V相空洞数", "W相空洞数", "模组位置"]
            missing = [x for x in need if x not in df.columns]
            if missing:
                LogUtil.append(f"❌ 汇总表缺少必要列：{missing}")
                return False
            wb = load_workbook(self.rel_path)
            ws = wb.active
            handle = skip_no_match = skip_pos_invalid = 0
            def safe_float(v):
                if pd.isna(v) or not str(v).strip():
                    return 0.0
                try:
                    return float(str(v).strip())
                except:
                    return 0.0
            for idx, row in df.iterrows():
                excel_row_no = idx + 2
                gz_raw = re.sub(r"\s+", "", str(row["工装编号"]))
                target = full_map.get(gz_raw)
                match_type = "精确匹配"
                if not target and len(gz_raw) >= 14:
                    target = prefix14_map.get(gz_raw[:14])
                    match_type = "前缀匹配"
                if not target:
                    skip_no_match += 1
                    LogUtil.append(f"  汇总表第{excel_row_no}行 | 工装号：{gz_raw} | 跳过：未匹配")
                    continue
                u, v, w = safe_float(row["U相空洞数"]), safe_float(row["V相空洞数"]), safe_float(row["W相空洞数"])
                pos_raw = str(row["模组位置"]).strip()
                try:
                    pos = int(float(pos_raw))
                except:
                    skip_pos_invalid += 1
                    continue
                if pos < 1 or pos > 8:
                    skip_pos_invalid += 1
                    continue
                colU, colV, colW = 2 + pos - 1, 10 + pos - 1, 18 + pos - 1
                def get_val(c):
                    try:
                        return float(ws.cell(target, c).value) if ws.cell(target, c).value else 0.0
                    except:
                        return 0.0
                u_before, v_before, w_before = get_val(colU), get_val(colV), get_val(colW)
                ws.cell(target, colU, u_before + u)
                ws.cell(target, colV, v_before + v)
                ws.cell(target, colW, w_before + w)
                handle += 1
                LogUtil.append(f"  汇总表第{excel_row_no}行 | 工装号：{gz_raw} | {match_type} | 位号：{pos} | U：{u_before}+{u}={u_before+u} | V：{v_before}+{v}={v_before+v} | W：{w_before}+{w}={w_before+w}")
            z_update_count = 0
            for r in range(1, max(ws.max_row, 660) + 1):
                if str(ws.cell(r, 1).value or "").strip().startswith("GZ"):
                    total = 0
                    for c in range(2, 26):
                        try:
                            total += float(ws.cell(r, c).value) if ws.cell(r, c).value else 0
                        except:
                            pass
                    ws.cell(r, 26, int(total))
                    z_update_count += 1
            save_p = self.rel_path.replace(".xlsx", "_匹配完成_累加版.xlsx")
            wb.save(save_p)
            wb.close()
            self.result_path = save_p
            LogUtil.append(f"✅ 工装匹配完成")
            LogUtil.append(f"  成功匹配并累加：{handle} 条")
            LogUtil.append(f"  未找到工装：{skip_no_match} 条")
            LogUtil.append(f"  模组位置无效：{skip_pos_invalid} 条")
            LogUtil.append(f"  Z列更新：{z_update_count} 条")
            return True
        except Exception as e:
            LogUtil.append(f"❌ 匹配异常：{e}")
            LogUtil.append(traceback.format_exc())
            return False
    def parse_visual_data(self):
        try:
            wb = load_workbook(self.result_path, data_only=True)
            ws = wb.active
            self.gz_total = {}
            self.position_total = [0] * 24
            for r in range(1, ws.max_row + 1):
                code = ws.cell(r, 1).value
                if not code or not str(code).strip().startswith("GZ"):
                    continue
                code = re.sub(r"\s+", "", str(code).strip())
                z = ws.cell(r, 26).value
                self.gz_total[code] = int(float(z)) if z else 0
                for i in range(24):
                    try:
                        self.position_total[i] += int(float(ws.cell(r, i + 2).value)) if ws.cell(r, i + 2).value else 0
                    except:
                        pass
            wb.close()
            return True
        except:
            return False
    def show_window(self, root_win):
        if not self.result_path or not self.parse_visual_data():
            messagebox.showerror("错误", "数据解析失败，无法打开可视化窗口", parent=root_win)
            return
        win = tk.Toplevel(root_win)
        win.title("工装空洞数据可视化")
        win.geometry("1100x700")
        win.resizable(True, True)
        bar = ttk.Frame(win, padding=15)
        bar.pack(fill=tk.X)
        fig = Figure(figsize=(11, 6), dpi=100)
        ax = fig.add_subplot(111)
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        ttk.Button(bar, text="按工装统计", command=lambda: self._switch("gz", canvas, fig)).pack(side=tk.LEFT, padx=10)
        ttk.Button(bar, text="按位号统计", command=lambda: self._switch("pos", canvas, fig)).pack(side=tk.LEFT, padx=10)
        self.ctrl_frame = ttk.Frame(bar)
        self.ctrl_frame.pack(side=tk.LEFT, padx=30)
        ttk.Label(self.ctrl_frame, text="展示数量：").pack(side=tk.LEFT, padx=5)
        self.show_n = tk.StringVar(value="20")
        cb = ttk.Combobox(self.ctrl_frame, textvariable=self.show_n, state="readonly", width=8)
        cb["values"] = ("10", "20", "30", "50", "全部")
        cb.pack(side=tk.LEFT, padx=5)
        ttk.Button(self.ctrl_frame, text="刷新", command=lambda: self._draw_gz(fig, ax)).pack(side=tk.LEFT, padx=10)
        self.current = "gz"
        self._draw_gz(fig, ax)
        canvas.draw()
    def _switch(self, typ, canvas, fig):
        self.current = typ
        fig.clear()
        ax = fig.add_subplot(111)
        if typ == "gz":
            self.ctrl_frame.pack(side=tk.LEFT, padx=30)
            self._draw_gz(fig, ax)
        else:
            self.ctrl_frame.pack_forget()
            self._draw_pos(fig, ax)
        canvas.draw()
    def _draw_gz(self, fig, ax):
        ax.clear()
        lst = sorted(self.gz_total.items(), key=lambda x: x[1], reverse=True)
        n = self.show_n.get()
        show = lst if n == "全部" else lst[:int(n)]
        codes, vals = [x[0] for x in show], [x[1] for x in show]
        y = range(len(codes))
        ax.barh(y, vals, height=0.7)
        ax.set_yticks(y)
        ax.set_yticklabels(codes, fontsize=8)
        ax.invert_yaxis()
        ax.set_title("工装总空洞数排名")
        ax.set_xlabel("总空洞数（个）")
        ax.grid(axis="x", linestyle="--", alpha=0.3)
        ax.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
        mx = max(vals) if vals else 1
        for i, v in enumerate(vals):
            ax.text(v + mx * 0.01, i, f"{v}", va="center", fontsize=8)
        fig.tight_layout()
    def _draw_pos(self, fig, ax):
        ax.clear()
        u, v, w = self.position_total[0:8], self.position_total[8:16], self.position_total[16:24]
        u_sum, v_sum, w_sum = sum(u), sum(v), sum(w)
        x = np.arange(8)
        bw, rs = 0.25, 0.06
        data_list, labels = [u, v, w], ["U相", "V相", "W相"]
        for pi in range(3):
            data = data_list[pi]
            for i in range(8):
                h = data[i]
                if h <= 0:
                    continue
                left = x[i] + (pi - 1) * bw - bw / 2
                ax.add_patch(FancyBboxPatch((left, 0), bw, h, boxstyle=f"round,pad=0,rounding_size={rs}", edgecolor="none"))
                total_max = max(u + v + w) if (u + v + w) else 1
                ax.text(left + bw / 2, h + total_max * 0.01, f"{int(h)}", ha="center", va="bottom", fontsize=8)
        ax.set_xlim(-0.8, 7.8)
        total_max = max(u + v + w) if (u + v + w) else 1
        ax.set_ylim(0, total_max * 1.15)
        ax.set_xticks(x)
        ax.set_xticklabels([str(i) for i in range(1, 9)])
        ax.set_title("各炉位总空洞数统计")
        ax.set_xlabel("在炉位号")
        ax.set_ylabel("总空洞数（个）")
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
        ax.legend(handles=[Patch(label=labels[i]) for i in range(3)], loc="upper right", fontsize=9)
        items = [f"U相总空洞：{u_sum}", f"V相总空洞：{v_sum}", f"W相总空洞：{w_sum}"]
        for i, txt in enumerate(items):
            ax.text(0.18 + i * 0.32, -0.18, txt, transform=ax.transAxes, ha="center", va="top", fontsize=10, bbox=dict(boxstyle="square,pad=0.6", fc="white", lw=1.5))
        fig.subplots_adjust(bottom=0.2)

class MainApp:
    def __init__(self, root):
        self.root = root
        root.title("X-Ray工装空洞分析全流程工具（批量版）")
        root.geometry("850x600")
        root.resizable(True, True)
        root.configure(bg=AppConfig.BG_COLOR)
        root.option_add("*Font", AppConfig.NORMAL_FONT)
        self.var_step1_root, self.var_step2_dir, self.var_step3_rel = tk.StringVar(), tk.StringVar(), tk.StringVar()
        self.batch_results, self.merged_file, self.matcher = [], "", None
        self._build_ui()
        LogUtil.bind(self.log_text, root)
        LogUtil.append("程序启动完成")
        LogUtil.append("步骤1：选择总根目录，程序自动识别MD/MN/DD任务令")
        LogUtil.append("任务令规则：MD/MN开头11位；DD开头12位")
    def _build_ui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill=tk.BOTH, expand=True)
        step1 = ttk.LabelFrame(main, text=" 步骤1：批量处理任务令 ", padding=8)
        step1.pack(fill=tk.X, pady=4)
        step1.columnconfigure(1, weight=1)
        ttk.Label(step1, text="总根目录：").grid(row=0, column=0, sticky="w", padx=4, pady=3)
        ttk.Entry(step1, textvariable=self.var_step1_root).grid(row=0, column=1, sticky="ew", padx=4, pady=3)
        ttk.Button(step1, text="浏览", width=8, command=self._browse_step1).grid(row=0, column=2, padx=4, pady=3)
        ttk.Button(step1, text="启动批量全自动任务", width=20, command=self._run_step1).grid(row=0, column=3, padx=8, pady=3)
        step2 = ttk.LabelFrame(main, text=" 步骤2：多任务令汇总表合并 ", padding=8)
        step2.pack(fill=tk.X, pady=4)
        step2.columnconfigure(1, weight=1)
        ttk.Label(step2, text="汇总表目录：").grid(row=0, column=0, sticky="w", padx=4, pady=3)
        ttk.Entry(step2, textvariable=self.var_step2_dir).grid(row=0, column=1, sticky="ew", padx=4, pady=3)
        ttk.Button(step2, text="浏览", width=8, command=self._browse_step2).grid(row=0, column=2, padx=4, pady=3)
        ttk.Button(step2, text="执行多表合并", width=20, command=self._run_step2).grid(row=0, column=3, padx=8, pady=3)
        step3 = ttk.LabelFrame(main, text=" 步骤3：工装空洞关系匹配与可视化 ", padding=8)
        step3.pack(fill=tk.X, pady=4)
        step3.columnconfigure(1, weight=1)
        ttk.Label(step3, text="工装及在炉位置记录表：").grid(row=0, column=0, sticky="w", padx=4, pady=3)
        ttk.Entry(step3, textvariable=self.var_step3_rel).grid(row=0, column=1, sticky="ew", padx=4, pady=3)
        ttk.Button(step3, text="浏览", width=8, command=self._browse_step3).grid(row=0, column=2, padx=4, pady=3)
        ttk.Button(step3, text="执行匹配及可视化展示", width=20, command=self._run_step3_match).grid(row=0, column=3, padx=8, pady=3)
        ttk.Label(main, text="执行日志", font=AppConfig.BOLD_FONT).pack(anchor="w", pady=(6, 2))
        log_wrap = ttk.Frame(main, relief=tk.GROOVE, padding=6)
        log_wrap.pack(fill=tk.BOTH, expand=True)
        self.log_text = scrolledtext.ScrolledText(log_wrap, font=AppConfig.LOG_FONT, relief=tk.FLAT)
        self.log_text.pack(fill=tk.BOTH, expand=True)
    def _browse_step1(self):
        p = filedialog.askdirectory(title="选择总根目录")
        if p:
            self.var_step1_root.set(p)
            LogUtil.append(f"已选择步骤1根目录：{p}")
    def _browse_step2(self):
        p = filedialog.askdirectory(title="选择汇总表所在目录")
        if p:
            self.var_step2_dir.set(p)
            LogUtil.append(f"已选择步骤2汇总表目录：{p}")
    def _browse_step3(self):
        p = filedialog.askopenfilename(title="选择工装及在炉位置记录表", filetypes=[("Excel文件", "*.xlsx")])
        if p:
            self.var_step3_rel.set(p)
            LogUtil.append(f"已选择步骤3工装记录表：{p}")
    def _run_step1(self):
        p = self.var_step1_root.get().strip()
        if not p or not os.path.isdir(p):
            messagebox.showwarning("提示", "请先选择有效的总根目录")
            return
        def _finish(results):
            self.batch_results = results
            self.root.after(0, lambda: messagebox.showinfo("完成", "批量处理完成！\n新生成汇总表已统一归档到「图片匹配结果汇总」文件夹"))
        def _thread():
            runner = BatchTaskRunner(p)
            if runner.scan_tasks() == 0:
                LogUtil.append("❌ 总根目录下一层未找到符合条件的任务令文件夹")
                LogUtil.append("要求：MD/MN开头11位；DD开头12位")
                _finish([])
                return
            runner.run_all(_finish)
        threading.Thread(target=_thread, daemon=True).start()
    def _run_step2(self):
        dir_p = self.var_step2_dir.get().strip()
        if not dir_p or not os.path.isdir(dir_p):
            messagebox.showwarning("提示", "请先选择汇总表所在目录")
            return
        target_files = [os.path.join(dir_p, f) for f in os.listdir(dir_p) if os.path.isfile(os.path.join(dir_p, f)) and f.lower().endswith(".xlsx") and "_图片匹配汇总结果" in f and not f.startswith("~$")]
        if not target_files:
            messagebox.showwarning("提示", "该目录下未找到包含「_图片匹配汇总结果」的Excel文件")
            return
        res = SummaryMerger(os.path.dirname(dir_p)).run(sorted(target_files))
        if res:
            self.merged_file = res
            messagebox.showinfo("完成", f"合并完成！\n总表文件：{os.path.basename(res)}")
    def _run_step3_match(self):
        rel_p = self.var_step3_rel.get().strip()
        if not self.merged_file:
            messagebox.showwarning("提示", "请先执行步骤2生成合并总表")
            return
        if not rel_p or not os.path.isfile(rel_p):
            messagebox.showwarning("提示", "请选择工装及在炉位置记录表")
            return
        self.matcher = VoidMatcher(self.merged_file, rel_p)
        if self.matcher.run():
            messagebox.showinfo("完成", "匹配完成！即将打开可视化窗口")
            self.matcher.show_window(self.root)

if __name__ == "__main__":
    root = tk.Tk()
    MainApp(root)
    root.mainloop()
